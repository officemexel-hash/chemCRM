import csv
import io
import logging

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.base import new_uuid
from app.db.models.bulk_import import BulkImportItem, BulkImportJob
from app.db.models.substance import Substance
from app.services.audit_log import AuditLogService
from app.services.cas_validator import is_valid_cas
from app.services.substance_enrichment import SubstanceEnrichmentService

logger = logging.getLogger(__name__)


class BulkImportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.audit = AuditLogService(db)

    def create_job(self, filename: str, original_filename: str | None, created_by: str | None = None) -> BulkImportJob:
        job = BulkImportJob(
            id=new_uuid(),
            filename=filename,
            original_filename=original_filename,
            status="pending",
            created_by=created_by,
        )
        self.db.add(job)
        self.db.flush()
        self.audit.log("bulk_import_job_created", "bulk_import_job", job.id, {"original_filename": original_filename})
        return job

    def parse_csv(self, content: bytes, job_id: str) -> list[BulkImportItem]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        items: list[BulkImportItem] = []
        for row_num, row in enumerate(reader, start=2):
            cas_raw = row.get("cas") or row.get("CAS") or row.get("Cas") or row.get("CAS_Number", "").strip()
            if not cas_raw:
                continue
            item = BulkImportItem(
                id=new_uuid(),
                job_id=job_id,
                row_number=row_num,
                cas_raw=cas_raw.strip(),
                cas_valid=is_valid_cas(cas_raw.strip()),
                status="pending",
            )
            items.append(item)
        return items

    def parse_excel(self, content: bytes, job_id: str) -> list[BulkImportItem]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            return []

        # Find CAS column
        header_row = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        cas_col = None
        for idx, header in enumerate(header_row):
            if header in ("cas", "cas_number", "cas no", "cas no.", "cas#"):
                cas_col = idx
                break
        if cas_col is None:
            # Fallback: use first column
            cas_col = 0

        items: list[BulkImportItem] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if cas_col >= len(row) or not row[cas_col]:
                continue
            cas_raw = str(row[cas_col]).strip()
            if not cas_raw:
                continue
            item = BulkImportItem(
                id=new_uuid(),
                job_id=job_id,
                row_number=row_num,
                cas_raw=cas_raw,
                cas_valid=is_valid_cas(cas_raw),
                status="pending",
            )
            items.append(item)
        wb.close()
        return items

    def validate_and_create(self, job_id: str) -> BulkImportJob:
        job = self.db.get(BulkImportJob, job_id)
        if job is None:
            raise ValueError("Job not found")
        job.status = "processing"
        self.db.flush()

        items = list(self.db.scalars(select(BulkImportItem).where(BulkImportItem.job_id == job_id)))
        valid = invalid = 0

        for item in items:
            if not item.cas_valid:
                item.status = "invalid"
                item.error_message = f"Invalid CAS checksum: {item.cas_raw}"
                invalid += 1
                continue

            # Check if substance already exists
            existing = self.db.scalar(select(Substance).where(Substance.cas == item.cas_raw))
            if existing:
                item.status = "existing"
                item.substance_id = existing.id
                valid += 1
                continue

            # Create new substance
            substance = Substance(
                id=new_uuid(),
                cas=item.cas_raw,
                regulatory_status="unknown",
            )
            self.db.add(substance)
            self.db.flush()
            item.status = "created"
            item.substance_id = substance.id
            valid += 1

        job.valid_rows = valid
        job.invalid_rows = invalid
        job.total_rows = len(items)
        job.status = "completed"
        self.db.flush()
        self.audit.log(
            "bulk_import_completed",
            "bulk_import_job",
            job_id,
            {"total": job.total_rows, "valid": valid, "invalid": invalid},
        )
        self.db.commit()
        self.db.refresh(job)
        return job

    def enrich_substances(self, job_id: str) -> dict:
        """Enrich all newly created substances from a bulk import job."""
        items = list(
            self.db.scalars(
                select(BulkImportItem).where(
                    BulkImportItem.job_id == job_id,
                    BulkImportItem.status == "created",
                    BulkImportItem.substance_id.isnot(None),
                )
            )
        )
        enrichment_service = SubstanceEnrichmentService(self.db)
        enriched = 0
        for item in items:
            try:
                enrichment_service.enrich(item.substance_id)
                enriched += 1
            except Exception as e:
                logger.warning("Failed to enrich substance %s: %s", item.substance_id, e)
        self.db.commit()
        self.audit.log("bulk_enrich_completed", "bulk_import_job", job_id, {"enriched": enriched})
        return {"enriched": enriched, "total": len(items)}

    def get_job_status(self, job_id: str) -> BulkImportJob:
        job = self.db.get(BulkImportJob, job_id)
        if job is None:
            raise ValueError("Job not found")
        return job
