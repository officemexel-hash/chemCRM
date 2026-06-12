"""Report Generator — creates comparison PDF/Excel reports and supplier rankings."""

import io
import logging
import os
from datetime import datetime

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.base import new_uuid
from app.db.models.document import GeneratedDocument
from app.db.models.quote import Quote
from app.db.models.substance import Substance
from app.db.models.supplier import SupplierCompany
from app.db.models.campaign import RfqCampaign

logger = logging.getLogger(__name__)


class ReportGeneratorService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def generate_comparison_pdf(self, campaign_id: str) -> GeneratedDocument:
        """Generate a PDF comparison report for all quotes in a campaign."""
        campaign, substance, quotes = self._load_campaign_quotes(campaign_id)
        suppliers = self._load_suppliers_for_quotes(quotes)

        html = self._build_comparison_html(campaign, substance, quotes, suppliers)
        return self._save_report("report_comparison", html, campaign_id=campaign_id, substance_id=substance.id)

    def generate_comparison_excel(self, campaign_id: str) -> GeneratedDocument:
        """Generate an Excel comparison report for all quotes."""
        campaign, substance, quotes = self._load_campaign_quotes(campaign_id)
        suppliers = self._load_suppliers_for_quotes(quotes)

        workbook_bytes = self._build_excel(campaign, substance, quotes, suppliers)
        return self._save_excel_report("report_comparison", workbook_bytes, campaign_id=campaign_id, substance_id=substance.id)

    def generate_supplier_ranking(self, campaign_id: str) -> list[dict]:
        """Rank quotes by price, lead time, supplier quality, risk, and document completeness."""
        campaign, substance, quotes = self._load_campaign_quotes(campaign_id)
        suppliers = self._load_suppliers_for_quotes(quotes)

        scored: list[dict] = []
        for quote in quotes:
            supplier = suppliers.get(quote.company_id)
            scores = self._score_quote(quote, supplier)
            scored.append({
                "quote_id": quote.id,
                "supplier_name": supplier.name if supplier else "Unknown",
                "country": supplier.country if supplier else None,
                "price": str(quote.price) if quote.price else None,
                "currency": quote.currency,
                "lead_time": quote.lead_time,
                "incoterms": quote.incoterms,
                **scores,
            })

        # Sort by total score descending
        scored.sort(key=lambda x: x["total_score"], reverse=True)

        # Add rank and recommended flag
        for i, item in enumerate(scored):
            item["rank"] = i + 1
            item["recommended"] = i == 0 and item["total_score"] > 50

        return scored

    def _score_quote(self, quote: Quote, supplier: SupplierCompany | None) -> dict:
        """Score a quote across multiple dimensions (0-100 each)."""
        # Price score: lower is better (normalize relative to cheapest)
        price_score = 50.0
        if quote.price:
            try:
                price_val = float(quote.price)
                price_score = max(0, 100 - price_val * 0.1)  # Simplified scoring
            except (ValueError, TypeError):
                pass

        # Supplier quality score
        supplier_quality = 50.0
        if supplier:
            try:
                supplier_quality = min(100, float(supplier.supplier_score or 0) * 10)
            except (ValueError, TypeError):
                pass

        # Risk score: lower risk = higher score
        risk_score = 50.0
        if supplier:
            try:
                risk_score = max(0, 100 - float(supplier.risk_score or 0))
            except (ValueError, TypeError):
                pass

        # Document completeness
        docs = sum(1 for field in [quote.coa_available, quote.sds_available, quote.reach_status] if field)
        doc_score = (docs / 3) * 100

        total = price_score * 0.35 + supplier_quality * 0.20 + risk_score * 0.25 + doc_score * 0.20

        return {
            "total_score": round(total, 1),
            "price_score": round(price_score, 1),
            "supplier_quality_score": round(supplier_quality, 1),
            "risk_score": round(risk_score, 1),
            "document_completeness": round(doc_score, 1),
        }

    def _build_comparison_html(self, campaign, substance, quotes, suppliers) -> str:
        """Build HTML comparison table."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        rows_html = ""
        for q in quotes:
            supplier = suppliers.get(q.company_id)
            name = supplier.name if supplier else "Unknown"
            row = f"""
            <tr>
                <td>{name}</td>
                <td>{q.price or '-'} {q.currency or ''}</td>
                <td>{q.moq or '-'}</td>
                <td>{q.lead_time or '-'}</td>
                <td>{q.incoterms or '-'}</td>
                <td>{'✓' if q.coa_available else '✗'}</td>
                <td>{'✓' if q.sds_available else '✗'}</td>
                <td>{q.reach_status or '-'}</td>
                <td>{q.packaging or '-'}</td>
            </tr>"""
            rows_html += row

        return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 30px; color: #1a1a1a; }}
h1 {{ color: #1a3a5c; font-size: 18pt; }}
h2 {{ color: #334155; font-size: 13pt; margin-top: 20px; }}
table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 9pt; }}
th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; }}
th {{ background-color: #1a3a5c; color: white; }}
tr:nth-child(even) {{ background-color: #f8fafc; }}
.meta {{ color: #64748b; font-size: 9pt; margin-bottom: 15px; }}
</style></head><body>
<h1>Quote Comparison Report</h1>
<div class="meta">
  Campaign: {campaign.id[:8]} | Substance: {substance.primary_name or substance.cas} (CAS: {substance.cas}) | Generated: {now}
</div>
<h2>Side-by-Side Comparison</h2>
<table>
<tr><th>Supplier</th><th>Price</th><th>MOQ</th><th>Lead Time</th><th>Incoterms</th><th>COA</th><th>SDS</th><th>REACH</th><th>Packaging</th></tr>
{rows_html}
</table>
</body></html>"""

    def _build_excel(self, campaign, substance, quotes, suppliers) -> bytes:
        """Build Excel workbook with 3 sheets."""
        from openpyxl import Workbook

        wb = Workbook()

        # Sheet 1: Quotes
        ws1 = wb.active
        ws1.title = "Quotes"
        ws1.append(["Supplier", "Price", "Currency", "MOQ", "Lead Time", "Incoterms", "COA", "SDS", "REACH", "Packaging", "Payment Terms"])
        for q in quotes:
            supplier = suppliers.get(q.company_id)
            ws1.append([
                supplier.name if supplier else "Unknown",
                str(q.price) if q.price else "",
                q.currency or "",
                q.moq or "",
                q.lead_time or "",
                q.incoterms or "",
                "Yes" if q.coa_available else "No",
                "Yes" if q.sds_available else "No",
                q.reach_status or "",
                q.packaging or "",
                q.payment_terms or "",
            ])

        # Sheet 2: Scoring
        ws2 = wb.create_sheet("Scoring")
        ws2.append(["Rank", "Supplier", "Total Score", "Price Score", "Quality Score", "Risk Score", "Doc Score", "Recommended"])
        ranking = self.generate_supplier_ranking(campaign.id)
        for r in ranking:
            ws2.append([
                r["rank"],
                r["supplier_name"],
                r["total_score"],
                r["price_score"],
                r["supplier_quality_score"],
                r["risk_score"],
                r["document_completeness"],
                "★" if r["recommended"] else "",
            ])

        # Sheet 3: Summary
        ws3 = wb.create_sheet("Summary")
        ws3.append(["Campaign ID", campaign.id])
        ws3.append(["Substance", substance.primary_name or substance.cas])
        ws3.append(["CAS", substance.cas or ""])
        ws3.append(["Quantity", campaign.quantity or ""])
        ws3.append(["Total Quotes", len(quotes)])
        ws3.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _load_campaign_quotes(self, campaign_id: str):
        campaign = self.db.get(RfqCampaign, campaign_id)
        if not campaign:
            raise ValueError("Campaign not found")
        substance = self.db.get(Substance, campaign.substance_id)
        if not substance:
            raise ValueError("Substance not found")
        quotes = list(self.db.scalars(select(Quote).where(Quote.campaign_id == campaign_id)))
        return campaign, substance, quotes

    def _load_suppliers_for_quotes(self, quotes) -> dict:
        supplier_ids = list({q.company_id for q in quotes})
        suppliers = list(self.db.scalars(select(SupplierCompany).where(SupplierCompany.id.in_(supplier_ids))))
        return {s.id: s for s in suppliers}

    def _save_report(self, doc_type: str, html: str, campaign_id: str | None = None, substance_id: str | None = None) -> GeneratedDocument:
        settings = get_settings()
        os.makedirs(settings.storage_path, exist_ok=True)
        doc = GeneratedDocument(id=new_uuid(), doc_type=doc_type, campaign_id=campaign_id, substance_id=substance_id, status="generated")
        filename = f"{doc_type}_{doc.id[:8]}.pdf"
        filepath = os.path.join(settings.storage_path, filename)
        try:
            from weasyprint import HTML
            HTML(string=html).write_pdf(filepath)
            doc.file_path = filepath
            doc.file_size_bytes = os.path.getsize(filepath)
        except Exception:
            html_path = filepath.replace(".pdf", ".html")
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            doc.file_path = html_path
            doc.file_size_bytes = os.path.getsize(html_path)
        self.db.add(doc)
        self.db.commit()
        self.db.refresh(doc)
        return doc

    def _save_excel_report(self, doc_type: str, data: bytes, campaign_id: str | None = None, substance_id: str | None = None) -> GeneratedDocument:
        settings = get_settings()
        os.makedirs(settings.storage_path, exist_ok=True)
        doc = GeneratedDocument(id=new_uuid(), doc_type=doc_type, campaign_id=campaign_id, substance_id=substance_id, status="generated")
        filename = f"{doc_type}_{doc.id[:8]}.xlsx"
        filepath = os.path.join(settings.storage_path, filename)
        with open(filepath, "wb") as f:
            f.write(data)
        doc.file_path = filepath
        doc.file_size_bytes = len(data)
        self.db.add(doc)
        self.db.commit()
        self.db.refresh(doc)
        return doc
